from nluas.language.analyzer_proxy import *
import openpyxl
from openpyxl.styles import colors

from graphviz import Digraph

from collections import OrderedDict

from openpyxl.styles import PatternFill


def split_sentence(sentence):
	s = sentence.replace(".", " . ").replace(",", " , ")
	return s.split()



def match_spans2(split, spans):
	final = OrderedDict()
	for span in spans:
		lr = span['span']
		key = "{}[{}]".format(span['type'], str(span['id']))
		final[key] = (split[lr[0]:lr[1]], lr, span['id'])
	return final


def get_roles(item, parse):
	roles = []
	for i in parse.__features__.values():
		for role, filler in i.__items__():
			if filler.index() == item.index() and role != "m":

				#role_frame = "{} ({})".format(role, i.type())
				roles.append(role)
	return roles

def get_semantics(cxn, ID, parse):
	for i in parse.__features__.values():
		for role, filler in i.__items__():
			if filler.index() == ID:
				if hasattr(filler, "m"):
					return [filler.m.type(), get_roles(filler.m, parse)]
	return None



def tag_excel(matched, split, parse):
	wb = openpyxl.load_workbook('example.xlsx')

	redFill = PatternFill(start_color=colors.RED, end_color=colors.RED, fill_type="solid")
	greenFill = PatternFill(start_color=colors.GREEN, end_color=colors.GREEN, fill_type="solid")

	sheet = wb.get_sheet_by_name("Sheet1")
	# Write words to file
	for index in range(len(split)):
		cell = sheet.cell(row=1, column=index+1)
		cell.set_explicit_value(split[index])
	# Write spans
	row = 2
	for key, value in matched.items():
		semantics = get_semantics(key, value[2], parse)
		for j in range(value[1][0], value[1][1]):


			cell = sheet.cell(row=row, column=j+1)
			cell.set_explicit_value(key)
			cell.fill = redFill
			
			
			if semantics:
				cell_s = sheet.cell(row=row+1, column=j+1)
				cell_s.set_explicit_value(semantics[0])
				#cell_s.set_explicit_value("{} :: {}".format(semantics[0], "+".join(semantics[1])))
				cell_s.fill = greenFill

		row+=2
	wb.save("example.xlsx")
	return sheet

def generate_graph(dot, parse, spans, observed=[], seen=[]):
	
	#s = matched[0]
	s = parse
	last_label = "{}[{}]".format(str(s.__type__), str(s.__index__))
	if last_label in spans:
		last_label += " '{}'".format(" ".join(spans[last_label][0]))
	dot.node(last_label, last_label)
	for key in parse.__fs__().__dict__.keys():
		new = getattr(s, key)

		new_label = "{}[{}]".format(str(new.__type__), str(new.__index__))
		if new_label in spans:
			new_label += " '{}'".format(" ".join(spans[new_label][0]))
		if new_label and str(new.__type__) != "None" and key != "features" and new.__type__ != "A123":
			if new.__typesystem__ == "SCHEMA":
				dot.node(new_label, new_label, color="red")
			elif new.__typesystem__ == "CONSTRUCTION":
				dot.node(new_label, new_label, color="green")
			elif new.__typesystem__ == "ONTOLOGY":
				dot.node(new_label, new_label, color="blue", shape="hexagon")
			new_key = "{}[{}]".format(key, new.__index__)
			dot.edge(last_label, new_label, label=key)
			if new.has_filler() and not new.index() in seen and not new_label in observed:# and new.__typesystem__ == "CONSTRUCTION":
				observed.append(new_label)
				seen.append(new.index())
				g = generate_graph(Digraph(), new, spans, observed, seen)
				dot.subgraph(g)
	return dot




analyzer = Analyzer("http://localhost:8090")


sentence = "Mary established the store"
split = split_sentence(sentence)

info = analyzer.full_parse(sentence)
parse, spans = info['parse'], info['spans']

s = match_spans2(split, spans[0])

dot = Digraph()
graph = generate_graph(dot, parse[0], s)
#graph.render('label_test/semspec.gv', view=True)


#sheet = tag_excel(s, split, parse[0])








